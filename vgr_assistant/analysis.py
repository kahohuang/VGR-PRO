from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .messaging import clean_text, generate_english_message
from .models import (
    STATUS_COOPERATING,
    STATUS_INITIAL,
    STATUS_LOST,
    STATUS_OVERDUE,
    STATUS_REPLIED,
    STATUS_SAMPLE_DELIVERED,
    STATUS_SAMPLE_SENT,
    STATUS_SECOND_TALKING,
    STATUS_SECOND_TOUCH,
    STATUS_VIDEO_POSTED,
    CreatorRecord,
)


def parse_video_count(value: Any) -> int:
    if value in (None, "", " "):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def has_whatsapp(record: CreatorRecord) -> bool:
    status = clean_text(record.whatsapp_status)
    contact = clean_text(record.whatsapp_contact)
    return status == "是" or (contact != "" and status != "已发短信")


def is_fathers_day(record: CreatorRecord) -> bool:
    return "父亲节" in clean_text(record.fathers_day_status)


def is_repeat_creator(record: CreatorRecord) -> bool:
    return parse_video_count(record.posted_video_count) > 0


def classify_authorization(value: Any) -> str:
    text = clean_text(value)
    if text == "":
        return "无授权"
    if "账号已授权" in text:
        return "账号已授权"
    if text == "否":
        return "无授权"
    if text == "是" or text.startswith("#"):
        return "ad code"
    return "ad code"


def parse_record_date(value: datetime | date | None) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def infer_stage(record: CreatorRecord) -> tuple[str, str, str]:
    coop_status = clean_text(record.cooperation_status)
    fathers_day_status = clean_text(record.fathers_day_status)
    authorization_type = classify_authorization(record.authorization_status)
    repeat_creator = is_repeat_creator(record)
    whatsapp_added = has_whatsapp(record)

    if "拒绝" in coop_status or "不符合" in coop_status:
        return "已关闭", "达人不适合继续推进", "停止跟进"

    if coop_status == STATUS_LOST:
        return "待二次唤醒", "回过一句后失联", "最后一次唤醒"

    if coop_status == STATUS_OVERDUE or fathers_day_status == "逾期":
        return "逾期待处理", "合作推进超期", "确认是否继续合作"

    video_live = (
        coop_status == STATUS_VIDEO_POSTED
        or "已发布视频" in fathers_day_status
        or parse_video_count(record.posted_video_count) > 0 and authorization_type != "无授权"
    )
    if video_live or coop_status == STATUS_VIDEO_POSTED:
        if authorization_type == "账号已授权":
            return "已完成账号授权", "账号已授权，可长期复用", "沉淀为长期投流资产"
        if authorization_type == "ad code":
            return "已拿到ad code", "单条视频已可投流", "确认是否继续争取账号授权"
        return "待获取ad code", "视频已发布但未给授权", "催ad code"

    if repeat_creator and coop_status == STATUS_COOPERATING:
        return "老达人复合作", "历史已发过视频，本轮无需寄样", "确认新视频排期"

    if coop_status == STATUS_SAMPLE_DELIVERED:
        return "已收货待发布", "已收货但迟迟不发视频", "催发布"

    if coop_status == STATUS_SAMPLE_SENT:
        return "已寄样待确认收货", "已寄样但未确认收货", "确认是否收货"

    if coop_status == STATUS_COOPERATING:
        return "待申请样品", "已同意合作但未申请样品", "提醒申请样品"

    if coop_status == STATUS_SECOND_TALKING:
        if whatsapp_added:
            return "WhatsApp沟通中", "已加WhatsApp，待推进合作细节", "推进合作细节"
        return "已回复待转WhatsApp", "已回复但未转到WhatsApp", "引导加WhatsApp"

    if coop_status == STATUS_REPLIED:
        if whatsapp_added:
            return "WhatsApp已接通", "已加WhatsApp，等待推进", "推进合作细节"
        return "已回复待转WhatsApp", "回过一句后未继续推进", "引导加WhatsApp"

    if coop_status == STATUS_SECOND_TOUCH:
        if whatsapp_added:
            return "已加WhatsApp待推进", "已加WhatsApp但未进入合作确认", "确认合作意向"
        return "二次建联待回复", "二次建联后仍未转化回复", "发送二次跟进"

    if coop_status == STATUS_INITIAL:
        return "初步建联待回复", "初次触达后等待回复", "保持观察或二次触达"

    if whatsapp_added:
        return "已加WhatsApp待推进", "已建立站外联系，待推进下一步", "确认合作意向"

    return "待人工判断", "现有字段不足以自动判断", "人工复核"


def compute_priority(stage: str, record: CreatorRecord) -> str:
    fathers_day = is_fathers_day(record)
    whatsapp_added = has_whatsapp(record)
    repeat_creator = is_repeat_creator(record)

    if stage in {"已关闭", "已完成账号授权", "已拿到ad code"}:
        return "低"
    if stage in {"待获取ad code", "已收货待发布", "待申请样品"}:
        return "高"
    if whatsapp_added or fathers_day or repeat_creator:
        return "高"
    if stage in {"已回复待转WhatsApp", "WhatsApp沟通中", "待二次唤醒", "逾期待处理"}:
        return "中"
    return "低"


def default_message_intent(result: dict[str, Any]) -> str:
    action = result["next_action"]
    mapping = {
        "提醒申请样品": "提醒他尽快申请样品，语气轻一点，顺便说我们会尽快审核",
        "确认是否收货": "问他有没有收到样品，如果收到了请告诉我，语气自然一点",
        "催发布": "提醒他收到样品后尽快发视频，顺便说我们会给流量支持，别太硬",
        "催ad code": "他发视频了，帮我催ad code，别显得太急",
        "确认新视频排期": "这个是老达人，不用提样品了，问他下周能不能发视频",
        "引导加WhatsApp": "请他加WhatsApp，方便确认合作细节，语气轻松一点",
        "确认合作意向": "问他这次合作的意向和时间安排，语气自然一点",
        "推进合作细节": "继续推进合作细节，确认视频内容和时间安排",
        "发送二次跟进": "二次跟进一下，问他有没有兴趣继续聊合作",
        "最后一次唤醒": "最后跟进一次，语气客气一点，看看他还愿不愿意合作",
        "确认是否继续合作": "这个达人拖期了，问他还要不要继续合作，语气直接一点",
        "沉淀为长期投流资产": "告诉他账号授权已经完成，感谢配合并保持长期联系",
        "确认是否继续争取账号授权": "已经拿到 ad code，顺便看看能不能争取账号授权",
    }
    return mapping.get(action, "帮我跟进这个达人，语气自然一点")


def analyze_records(records: list[CreatorRecord], today: date | None = None) -> list[dict[str, Any]]:
    today = today or date.today()
    results: list[dict[str, Any]] = []

    for record in records:
        stage, blocker, next_action = infer_stage(record)
        record_date = parse_record_date(record.date)
        stalled_days = (today - record_date).days if record_date else None
        result = {
            "date": record_date.isoformat() if record_date else "",
            "username": clean_text(record.username),
            "product": clean_text(record.product),
            "has_whatsapp": has_whatsapp(record),
            "whatsapp_status": clean_text(record.whatsapp_status),
            "whatsapp_contact": clean_text(record.whatsapp_contact),
            "posted_video_count": parse_video_count(record.posted_video_count),
            "tk_profile_url": clean_text(record.tk_profile_url),
            "cooperation_status": clean_text(record.cooperation_status),
            "owner": clean_text(record.owner),
            "authorization_raw": clean_text(record.authorization_status),
            "authorization_type": classify_authorization(record.authorization_status),
            "fathers_day_status": clean_text(record.fathers_day_status),
            "is_fathers_day": is_fathers_day(record),
            "notes": clean_text(record.notes),
            "is_repeat_creator": is_repeat_creator(record),
            "needs_sample": not is_repeat_creator(record)
            and stage in {"待申请样品", "已寄样待确认收货", "已收货待发布"},
            "stage": stage,
            "stalled_days": stalled_days if stalled_days is not None else "",
            "blocker": blocker,
            "next_action": next_action,
        }
        result["priority"] = compute_priority(stage, record)
        result["english_message"] = generate_english_message(
            default_message_intent(result),
            {
                "is_repeat_creator": result["is_repeat_creator"],
                "has_whatsapp": result["has_whatsapp"],
                "is_fathers_day": result["is_fathers_day"],
                "stage": result["stage"],
            },
        )
        results.append(result)

    priority_order = {"高": 0, "中": 1, "低": 2}
    results.sort(
        key=lambda item: (
            priority_order.get(item["priority"], 9),
            -int(item["stalled_days"] or 0),
            item["username"],
        )
    )
    return results


def load_creator_records(workbook_path: Path) -> list[CreatorRecord]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(min_row=2, values_only=True)
    records = []
    for row in rows:
        if not row or not clean_text(row[1]):
            continue
        records.append(
            CreatorRecord(
                date=row[0],
                username=clean_text(row[1]),
                product=clean_text(row[2]),
                whatsapp_status=clean_text(row[3]),
                whatsapp_contact=clean_text(row[4]),
                posted_video_count=row[5],
                tk_profile_url=clean_text(row[6]),
                cooperation_status=clean_text(row[9]),
                owner=clean_text(row[10]),
                authorization_status=clean_text(row[11]),
                fathers_day_status=clean_text(row[12]),
                notes=clean_text(row[13]),
            )
        )
    return records
